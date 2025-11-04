# -*- coding: utf-8 -*-
"""
Boca Velocity – populate DGR Sales Velocity from 3 Salesforce reports
- Handles nested groupings (Week > Owner)
- Extracts Last Week, Prior Week, and 6W Average per owner
- FIXED: Date logic now correctly handles LAST_WEEK throughout the week
"""

import os, re, time, jwt, json, requests
from datetime import datetime, date, timedelta
from pathlib import Path
from openpyxl import load_workbook
from collections import defaultdict
from html_template_base import generate_html_document, generate_html_table, save_json_history, load_json_history

# ---------- Paths / sheet ----------
HOME          = Path.home()
TEMPLATE_DIR  = HOME / "Desktop" / "My Weekly Reports - Python"
TEMPLATE_FILE = TEMPLATE_DIR / "BocaTemplate.xlsx"
SHEET_NAME    = "DGR Sales Velocity"

COL_REP         = "A"
MEET_LAST_COL   = "B"
MEET_PRIOR_COL  = "C"
MEET_6W_COL     = "E"
OPP_LAST_COL    = "G"
OPP_PRIOR_COL   = "H"
OPP_6W_COL      = "J"
CW_LAST_COL     = "L"
CW_PRIOR_COL    = "M"

NOTE_HINT_MEET  = "Boca Meeting Attended"
NOTE_HINT_OPP   = "Boca Sourced Opp Created Weekly"
NOTE_HINT_CW    = "Boca Sourced Weekly CW"

# OPTIONAL: Hard-code report IDs here to avoid needing cell notes
# Set to None to use notes instead
HARDCODED_REPORT_IDS = {
    "meet": "00OPQ0000077pUn2AI",
    "opp": "00OPQ0000077r1y2AA",
    "cw": "00OPQ0000077rMv2AI"
}

# ---------- SF config - JWT Authentication ----------
# This script uses JWT (JSON Web Token) Bearer authentication to connect to Salesforce.
# 
# Required Configuration (in sf_config.py - NOT committed to GitHub):
# - SF_USERNAME: Your Salesforce username
# - SF_CONSUMER_KEY: Connected App Consumer Key
# - SF_DOMAIN: 'login' for production, 'test' for sandbox
# - PRIVATE_KEY_FILE: Path to your RSA private key file (.pem)
# 
# See sf_config.py.example for template and setup instructions.
# Security Note: sf_config.py is gitignored and will NOT be committed to GitHub
#
import sf_config

def get_jwt_token():
    """
    Generate JWT token for Salesforce authentication using RS256 algorithm.
    
    Returns:
        dict: Token response containing 'access_token' and 'instance_url'
    
    Raises:
        requests.HTTPError: If authentication fails
    """
    # Read the RSA private key file
    with open(sf_config.PRIVATE_KEY_FILE, "r") as f:
        private_key = f.read()
    
    # Build JWT claim (payload) with required Salesforce fields
    claim = {
        "iss": sf_config.SF_CONSUMER_KEY,  # Issuer: Your Connected App Consumer Key
        "sub": sf_config.SF_USERNAME,      # Subject: Your Salesforce username
        "aud": f"https://{sf_config.SF_DOMAIN}.salesforce.com",  # Audience: Salesforce login URL
        "exp": int(time.time()) + 300,      # Expiration: 5 minutes from now
    }
    
    # Encode the JWT token using RS256 algorithm (RSA signature)
    assertion = jwt.encode(claim, private_key, algorithm="RS256")
    
    # Exchange JWT assertion for access token
    token_url = f"https://{sf_config.SF_DOMAIN}.salesforce.com/services/oauth2/token"
    r = requests.post(token_url, data={
        "grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",  # JWT Bearer grant type
        "assertion": assertion,  # The JWT token we just created
    })
    r.raise_for_status()
    return r.json()  # Returns {access_token, instance_url, etc.}

def api_base(tok, version="v61.0"):
    return f"{tok['instance_url']}/services/data/{version}"

def sf_headers(tok):
    return {"Authorization": f"Bearer {tok['access_token']}",
            "Content-Type": "application/json"}

# ---------- FIXED Date helpers (Sunday to Saturday weeks) ----------
def last_full_week(ref_date=None):
    """
    Returns LAST WEEK per Salesforce definition:
    Sunday-Saturday of the previous calendar week.
    
    Example: If today is Monday Oct 28, 2025:
    - This week started Sunday Oct 27
    - LAST WEEK = Sunday Oct 19 to Saturday Oct 25
    
    This matches Salesforce's LAST_WEEK date literal behavior.
    Works correctly regardless of what day of the week you run it.
    """
    if ref_date is None:
        ref_date = date.today()
    
    # Python weekday: Monday=0, Tuesday=1, ..., Sunday=6
    dow = ref_date.weekday()
    
    # Find this week's Sunday
    if dow == 6:  # Today is Sunday
        this_weeks_sunday = ref_date
    else:
        # Days since last Sunday
        days_since_sunday = (dow + 1) % 7
        this_weeks_sunday = ref_date - timedelta(days=days_since_sunday)
    
    # LAST WEEK ends the day before this week's Sunday
    last_weeks_saturday = this_weeks_sunday - timedelta(days=1)
    last_weeks_sunday = last_weeks_saturday - timedelta(days=6)
    
    return last_weeks_sunday, last_weeks_saturday

def prior_full_week(ref_date=None):
    """
    Returns the week before last_full_week.
    This is the week that ended 2 Saturdays ago.
    """
    lw_s, lw_e = last_full_week(ref_date)
    pw_e = lw_s - timedelta(days=1)  # Day before last week's Sunday
    pw_s = pw_e - timedelta(days=6)
    return pw_s, pw_e

# ---------- Analytics ----------
RID_RE = re.compile(r"/Report/(00O[a-zA-Z0-9]{12,18})/")

def scan_report_ids_from_notes(ws):
    found = {"meet": None, "opp": None, "cw": None}
    for row in ws.iter_rows(min_row=1, max_row=80, min_col=1, max_col=30):
        for cell in row:
            if cell.comment and cell.comment.text:
                m = RID_RE.search(cell.comment.text)
                if not m: continue
                rid = m.group(1)
                label = (cell.value or "")
                s = str(label).strip().lower()
                if NOTE_HINT_MEET.lower() in s: found["meet"] = rid
                elif NOTE_HINT_OPP.lower() in s: found["opp"] = rid
                elif NOTE_HINT_CW.lower() in s: found["cw"] = rid
    if not all(found.values()):
        ordered = []
        for row in ws.iter_rows(min_row=1, max_row=80, min_col=1, max_col=30):
            for cell in row:
                if cell.comment and cell.comment.text:
                    m = RID_RE.search(cell.comment.text)
                    if m: ordered.append(m.group(1))
        seen = []
        for r in ordered:
            if r not in seen: seen.append(r)
        if found["meet"] is None and len(seen) > 0: found["meet"] = seen[0]
        if found["opp"]  is None and len(seen) > 1: found["opp"]  = seen[1]
        if found["cw"]   is None and len(seen) > 2: found["cw"]   = seen[2]
    return found

def analytics_report_raw(tok, report_id):
    url = f"{api_base(tok)}/analytics/reports/{report_id}?includeDetails=true"
    r = requests.get(url, headers=sf_headers(tok))
    r.raise_for_status()
    return r.json()

def parse_week_range(week_str):
    """Parse '10/5/2025 - 10/11/2025' into (start_date, end_date)"""
    try:
        parts = week_str.split(" - ")
        start = datetime.strptime(parts[0].strip(), "%m/%d/%Y").date()
        end = datetime.strptime(parts[1].strip(), "%m/%d/%Y").date()
        return start, end
    except:
        return None, None

def extract_by_week_ranges(j, name, agg_idx, last_week_range, prior_week_range):
    """
    Extract data for specific week ranges only.
    Returns: last_map, prior_map, six_avg_map
    """
    fact_map = j.get("factMap", {})
    g_down = j.get("groupingsDown", {})
    
    # Build lookup structures
    week_data = {}  # week_label -> {owner -> value}
    owner_lookup = {}  # "week_idx_owner_idx" -> (week_label, owner_name)
    
    weeks = g_down.get("groupings", []) if g_down else []
    
    for week_idx, week_group in enumerate(weeks):
        week_label = week_group.get("label", "")
        owners = week_group.get("groupings", [])
        
        if week_label not in week_data:
            week_data[week_label] = {}
        
        for owner_idx, owner_group in enumerate(owners):
            owner_name = owner_group.get("label", "")
            key = f"{week_idx}_{owner_idx}"
            owner_lookup[key] = (week_label, owner_name)
    
    # Extract values from factMap
    for key, fact in fact_map.items():
        if "!T" not in key:
            continue
        
        parts = key.split("!")[0]  # "8_5"
        if parts not in owner_lookup:
            continue
        
        week_label, owner = owner_lookup[parts]
        aggs = fact.get("aggregates", [])
        
        if agg_idx < len(aggs):
            val = aggs[agg_idx].get("value")
            if val is not None:
                week_data[week_label][owner] = float(val)
    
    print(f"\n=== {name} (Agg index {agg_idx}) ===")
    print(f"Found {len(weeks)} weeks with data")
    
    # Identify which weeks match our target ranges
    last_week_label = None
    prior_week_label = None
    six_week_labels = []
    
    for week_label in week_data.keys():
        start, end = parse_week_range(week_label)
        if not start or not end:
            continue
        
        # Check if this is last week
        if start == last_week_range[0] and end == last_week_range[1]:
            last_week_label = week_label
        
        # Check if this is prior week
        if start == prior_week_range[0] and end == prior_week_range[1]:
            prior_week_label = week_label
        
        # Collect last 6 weeks for averaging
        if end <= last_week_range[1]:
            six_week_labels.append(week_label)
    
    # Keep only most recent 6 weeks for averaging
    six_week_labels = sorted(
        six_week_labels, 
        key=lambda w: parse_week_range(w)[1] if parse_week_range(w)[1] else date.min,
        reverse=True
    )[:6]
    
    print(f"Last week: {last_week_label}")
    print(f"Prior week: {prior_week_label}")
    print(f"6-week average period: {len(six_week_labels)} weeks")
    
    # Extract data
    last_map = week_data.get(last_week_label, {}) if last_week_label else {}
    prior_map = week_data.get(prior_week_label, {}) if prior_week_label else {}
    
    # Calculate 6-week average
    six_avg_map = {}
    all_owners = set()
    for wk in six_week_labels:
        all_owners.update(week_data.get(wk, {}).keys())
    
    for owner in all_owners:
        values = [week_data.get(wk, {}).get(owner, 0) for wk in six_week_labels]
        values = [v for v in values if v is not None]
        if values:
            six_avg_map[owner] = sum(values) / len(six_week_labels)
    
    print(f"Extracted: {len(last_map)} Last, {len(prior_map)} Prior, {len(six_avg_map)} 6W Avg")
    
    return last_map, prior_map, six_avg_map

# ---------- Excel ----------
def load_template_and_reps():
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"Missing template: {TEMPLATE_FILE}")
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb[SHEET_NAME]
    reps = []
    r = 6
    while True:
        v = ws[f"{COL_REP}{r}"].value
        if v is None: break
        s = str(v).strip()
        if s.lower().startswith("team average"): break
        reps.append((r, s))
        r += 1
    return wb, ws, reps

def write(ws, row, col, val):
    ws[f"{col}{row}"].value = val

# ---------- Main ----------
def main():
    print("=" * 60)
    print("BOCA SALES VELOCITY REPORT GENERATOR")
    print("=" * 60)
    
    print("\nLoading Boca template…")
    wb, ws, reps = load_template_and_reps()
    names = [n for _, n in reps]
    print(f"Found {len(reps)} reps: {names}")

    rids = scan_report_ids_from_notes(ws)
    
    # Override with hard-coded IDs if they exist
    if HARDCODED_REPORT_IDS.get("meet"):
        rids["meet"] = HARDCODED_REPORT_IDS["meet"]
    if HARDCODED_REPORT_IDS.get("opp"):
        rids["opp"] = HARDCODED_REPORT_IDS["opp"]
    if HARDCODED_REPORT_IDS.get("cw"):
        rids["cw"] = HARDCODED_REPORT_IDS["cw"]
    
    print(f"\nReports → Meetings:{rids['meet']}  Opp$:{rids['opp']}  CW$:{rids['cw']}")
    if not all(rids.values()):
        raise RuntimeError(f"Could not resolve 3 report IDs from notes. Found: {rids}")

    # Calculate week ranges using FIXED date logic
    today = date.today()
    last_wk = last_full_week(today)
    prior_wk = prior_full_week(today)
    
    print(f"\nToday: {today.strftime('%A, %B %d, %Y')}")
    print(f"\nTarget weeks:")
    print(f"  Last Week:  {last_wk[0]} to {last_wk[1]} ({(last_wk[1] - last_wk[0]).days + 1} days)")
    print(f"  Prior Week: {prior_wk[0]} to {prior_wk[1]} ({(prior_wk[1] - prior_wk[0]).days + 1} days)")
    print("=" * 60)

    tok = get_jwt_token()
    print("\n✓ Connected to Salesforce")

    # Fetch and extract with specific week targeting
    print("\nFetching reports...")
    
    meet_json = analytics_report_raw(tok, rids["meet"])
    meet_last, meet_prior, meet_6w = extract_by_week_ranges(
        meet_json, "Meetings", 0, last_wk, prior_wk
    )

    opp_json = analytics_report_raw(tok, rids["opp"])
    opp_last, opp_prior, opp_6w = extract_by_week_ranges(
        opp_json, "Opp $ (Last)", 0, last_wk, prior_wk
    )
    
    cw_json = analytics_report_raw(tok, rids["cw"])
    cw_last, cw_prior, cw_6w = extract_by_week_ranges(
        cw_json, "Closed Won $", 0, last_wk, prior_wk
    )

    # Prepare data for HTML and JSON
    dashboard_data = []
    json_data = {
        "period": {
            "last_week": {"start": str(last_wk[0]), "end": str(last_wk[1])},
            "prior_week": {"start": str(prior_wk[0]), "end": str(prior_wk[1])}
        },
        "data": {}
    }
    
    for row, rep in reps:
        # Gather all metrics
        meet_lw = meet_last.get(rep, 0)
        meet_pw = meet_prior.get(rep, 0)
        meet_6 = meet_6w.get(rep, 0)
        opp_lw = opp_last.get(rep, 0)
        opp_pw = opp_prior.get(rep, 0)
        opp_6 = opp_6w.get(rep, 0)
        cw_lw = cw_last.get(rep, 0)
        cw_pw = cw_prior.get(rep, 0)
        
        # Calculate WoW changes
        meet_wow = ((meet_lw - meet_pw) / meet_pw * 100) if meet_pw > 0 else 0
        opp_wow = ((opp_lw - opp_pw) / opp_pw * 100) if opp_pw > 0 else 0
        cw_wow = ((cw_lw - cw_pw) / cw_pw * 100) if cw_pw > 0 else 0
        
        # Store for JSON
        json_data["data"][rep] = {
            "meetings": {"last": meet_lw, "prior": meet_pw, "6w_avg": meet_6, "wow_pct": meet_wow},
            "opp_dollars": {"last": opp_lw, "prior": opp_pw, "6w_avg": opp_6, "wow_pct": opp_wow},
            "closed_won": {"last": cw_lw, "prior": cw_pw, "wow_pct": cw_wow}
        }
        
        # Prepare HTML row
        dashboard_data.append({
            "rep": rep,
            "meet_last": f"{meet_lw:.0f}",
            "meet_prior": f"{meet_pw:.0f}",
            "meet_6w": f"{meet_6:.1f}",
            "meet_wow": f"{meet_wow:+.1f}%" if meet_pw > 0 else "—",
            "opp_last": f"${opp_lw:,.0f}",
            "opp_prior": f"${opp_pw:,.0f}",
            "opp_6w": f"${opp_6:,.0f}",
            "opp_wow": f"{opp_wow:+.1f}%" if opp_pw > 0 else "—",
            "cw_last": f"${cw_lw:,.0f}",
            "cw_prior": f"${cw_pw:,.0f}",
            "cw_wow": f"{cw_wow:+.1f}%" if cw_pw > 0 else "—"
        })
    
    # Write Excel (optional - keep for backward compatibility)
    print("\n" + "=" * 60)
    print("WRITING TO EXCEL")
    print("=" * 60)
    
    for row, rep in reps:
        # Meetings - write value or 0 if not present
        write(ws, row, MEET_LAST_COL,  meet_last.get(rep, 0))
        write(ws, row, MEET_PRIOR_COL, meet_prior.get(rep, 0))
        write(ws, row, MEET_6W_COL,    meet_6w.get(rep, 0))

        # Opp $ - write value or 0 if not present
        write(ws, row, OPP_LAST_COL,   opp_last.get(rep, 0))
        write(ws, row, OPP_PRIOR_COL,  opp_prior.get(rep, 0))
        write(ws, row, OPP_6W_COL,     opp_6w.get(rep, 0))

        # CW $ - write value or 0 if not present
        write(ws, row, CW_LAST_COL,    cw_last.get(rep, 0))
        write(ws, row, CW_PRIOR_COL,   cw_prior.get(rep, 0))

        print(f"\n{rep}:")
        print(f"  Meet L/P/6W: {meet_last.get(rep, 0)} / {meet_prior.get(rep, 0)} / {meet_6w.get(rep, 0)}")
        print(f"  Opp$ L/P/6W: ${opp_last.get(rep, 0):,.0f} / ${opp_prior.get(rep, 0):,.0f} / ${opp_6w.get(rep, 0):,.0f}")
        print(f"  CW$  L/P:    ${cw_last.get(rep, 0):,.0f} / ${cw_prior.get(rep, 0):,.0f}")

    excel_out = TEMPLATE_DIR / f"{datetime.today().strftime('%Y-%m-%d')}-Boca Template.xlsx"
    wb.save(excel_out)
    print(f"\n✅ Excel saved: {excel_out}")
    
    # Generate HTML Dashboard
    print("\n" + "=" * 60)
    print("GENERATING HTML DASHBOARD")
    print("=" * 60)
    
    # Define table columns
    columns = [
        {"key": "rep", "label": "Rep", "sortable": True},
        {"key": "meet_last", "label": "Meetings<br>Last Week", "sortable": True, "align": "right"},
        {"key": "meet_prior", "label": "Meetings<br>Prior Week", "sortable": True, "align": "right"},
        {"key": "meet_6w", "label": "Meetings<br>6W Avg", "sortable": True, "align": "right"},
        {"key": "meet_wow", "label": "Meetings<br>WoW", "sortable": True, "align": "right"},
        {"key": "opp_last", "label": "Opp $<br>Last Week", "sortable": True, "align": "right"},
        {"key": "opp_prior", "label": "Opp $<br>Prior Week", "sortable": True, "align": "right"},
        {"key": "opp_6w", "label": "Opp $<br>6W Avg", "sortable": True, "align": "right"},
        {"key": "opp_wow", "label": "Opp $<br>WoW", "sortable": True, "align": "right"},
        {"key": "cw_last", "label": "CW $<br>Last Week", "sortable": True, "align": "right"},
        {"key": "cw_prior", "label": "CW $<br>Prior Week", "sortable": True, "align": "right"},
        {"key": "cw_wow", "label": "CW $<br>WoW", "sortable": True, "align": "right"}
    ]
    
    # Generate HTML table
    table_html = generate_html_table(columns, dashboard_data)
    
    # Generate complete HTML document
    title = "Boca Sales Velocity Report"
    subtitle = f"DGR Sales Velocity - Last Week: {last_wk[0]} to {last_wk[1]} | Prior Week: {prior_wk[0]} to {prior_wk[1]}"
    
    # Add JavaScript for sorting
    js_script = """
<script>
(function() {
  const tbody = document.querySelector('tbody');
  const headers = document.querySelectorAll('th.sort');
  let sortKey = 'rep';
  let sortDir = 'asc';
  
  function render() {
    const rows = Array.from(tbody.querySelectorAll('tr'));
    rows.sort((a, b) => {
      const aVal = a.querySelector(`td:nth-child(${getColumnIndex(sortKey)})`).textContent.trim();
      const bVal = b.querySelector(`td:nth-child(${getColumnIndex(sortKey)})`).textContent.trim();
      const aNum = parseFloat(aVal.replace(/[^0-9.-]/g, ''));
      const bNum = parseFloat(bVal.replace(/[^0-9.-]/g, ''));
      if (!isNaN(aNum) && !isNaN(bNum)) {
        return (aNum - bNum) * (sortDir === 'asc' ? 1 : -1);
      }
      return aVal.localeCompare(bVal) * (sortDir === 'asc' ? 1 : -1);
    });
    rows.forEach(r => tbody.appendChild(r));
  }
  
  function getColumnIndex(key) {
    const cols = ['rep', 'meet_last', 'meet_prior', 'meet_6w', 'meet_wow', 
                  'opp_last', 'opp_prior', 'opp_6w', 'opp_wow', 
                  'cw_last', 'cw_prior', 'cw_wow'];
    return cols.indexOf(key) + 1;
  }
  
  headers.forEach(th => {
    th.addEventListener('click', () => {
      sortKey = th.dataset.key;
      sortDir = sortDir === 'asc' ? 'desc' : 'asc';
      render();
    });
  });
})();
</script>
"""
    
    html_content = generate_html_document(title, subtitle, table_html, js_script)
    
    # Save HTML (in script directory)
    script_dir = Path(__file__).parent
    html_file = script_dir / "boca_velocity_latest.html"
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"✅ HTML Dashboard saved: {html_file}")
    
    # Save JSON History (week-specific: uses last_wk[0] as period start date)
    json_file = script_dir / "boca_velocity_history.json"
    history = save_json_history(json_file, json_data, last_wk[0], max_entries=12)
    week_key = last_wk[0].strftime("%Y-W%V") if isinstance(last_wk[0], date) else last_wk[0]
    print(f"✅ JSON History saved: {json_file} ({len(history)} entries) - Week: {week_key}")
    
    print("\n" + "=" * 60)
    print("✅ COMPLETE")
    print("=" * 60)
    print(f"\n📊 Outputs:")
    print(f"   Excel: {excel_out}")
    print(f"   HTML:  {html_file}")
    print(f"   JSON:  {json_file}")
    print(f"\n🌐 GitHub Pages URL:")
    print(f"   https://AstonAtInnovativeSol.github.io/salesforce-flow-metrics/boca_velocity_latest.html")

if __name__ == "__main__":
    main()
