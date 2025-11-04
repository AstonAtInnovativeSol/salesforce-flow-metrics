#!/usr/bin/env python3
"""
SOQL Backfill for JSON History - Populates initial JSON history from Salesforce data
For Boca Velocity and Pipeline Velocity reports

Usage:
    python backfill_json_history.py --start-date 2025-11-03 --weeks 12
    python backfill_json_history.py --start-date 2025-11-03  # defaults to 12 weeks
"""

import argparse
import sys
from datetime import date, timedelta
from pathlib import Path

# Import functions from both scripts
sys.path.insert(0, str(Path(__file__).parent))

# Import BocaSalesMotion2 functions
import BocaSalesMotion2 as boca
# Import pipev3 functions
import pipev3 as pipe
# Import HTML template functions
from html_template_base import save_json_history, get_week_key


def backfill_boca_velocity(start_date, weeks=12):
    """Backfill Boca Velocity JSON history from Salesforce"""
    print("\n" + "=" * 60)
    print("BACKFILLING BOCA VELOCITY JSON HISTORY")
    print("=" * 60)
    print(f"Start Date: {start_date}")
    print(f"Weeks: {weeks}")
    
    # Load template to get reps and report IDs
    from openpyxl import load_workbook
    wb = load_workbook(boca.TEMPLATE_FILE)
    ws = wb[boca.SHEET_NAME]
    reps = boca.load_template_and_reps()
    rep_names = [r[1] for r in reps]
    
    # Get report IDs
    rids = boca.scan_report_ids_from_notes(ws)
    if not all(rids.values()):
        print(f"⚠️  Warning: Could not resolve all report IDs. Found: {rids}")
        return
    
    print(f"\nReports: Meet={rids['meet']}, Opp={rids['opp']}, CW={rids['cw']}")
    
    # Authenticate
    print("\nAuthenticating...")
    tok = boca.get_jwt_token()
    print("✓ Connected")
    
    # Process each week
    script_dir = Path(__file__).parent
    json_file = script_dir / "boca_velocity_history.json"
    
    current_date = date.fromisoformat(start_date) if isinstance(start_date, str) else start_date
    
    print(f"\nProcessing {weeks} weeks starting from {current_date}...")
    
    for week_num in range(weeks):
        # Calculate week ranges for this iteration
        last_wk = boca.last_full_week(current_date)
        prior_wk = boca.prior_full_week(current_date)
        
        print(f"\nWeek {week_num + 1}: {last_wk[0]} to {last_wk[1]}")
        
        # Fetch reports for this week
        try:
            
            meet_json = boca.analytics_report_raw(tok, rids["meet"])
            meet_last, meet_prior, meet_6w = boca.extract_by_week_ranges(
                meet_json, "Meetings", 0, last_wk, prior_wk
            )
            
            opp_json = boca.analytics_report_raw(tok, rids["opp"])
            opp_last, opp_prior, opp_6w = boca.extract_by_week_ranges(
                opp_json, "Opp $ (Last)", 0, last_wk, prior_wk
            )
            
            cw_json = boca.analytics_report_raw(tok, rids["cw"])
            cw_last, cw_prior, cw_6w = boca.extract_by_week_ranges(
                cw_json, "Closed Won $", 0, last_wk, prior_wk
            )
            
            # Build JSON data structure
            json_data = {
                "period": {
                    "last_week": {"start": str(last_wk[0]), "end": str(last_wk[1])},
                    "prior_week": {"start": str(prior_wk[0]), "end": str(prior_wk[1])}
                },
                "data": {}
            }
            
            for rep in rep_names:
                meet_lw = meet_last.get(rep, 0)
                meet_pw = meet_prior.get(rep, 0)
                meet_6 = meet_6w.get(rep, 0)
                opp_lw = opp_last.get(rep, 0)
                opp_pw = opp_prior.get(rep, 0)
                opp_6 = opp_6w.get(rep, 0)
                cw_lw = cw_last.get(rep, 0)
                cw_pw = cw_prior.get(rep, 0)
                
                meet_wow = ((meet_lw - meet_pw) / meet_pw * 100) if meet_pw > 0 else 0
                opp_wow = ((opp_lw - opp_pw) / opp_pw * 100) if opp_pw > 0 else 0
                cw_wow = ((cw_lw - cw_pw) / cw_pw * 100) if cw_pw > 0 else 0
                
                json_data["data"][rep] = {
                    "meetings": {"last": meet_lw, "prior": meet_pw, "6w_avg": meet_6, "wow_pct": meet_wow},
                    "opp_dollars": {"last": opp_lw, "prior": opp_pw, "6w_avg": opp_6, "wow_pct": opp_wow},
                    "closed_won": {"last": cw_lw, "prior": cw_pw, "wow_pct": cw_wow}
                }
            
            # Save to history (week-specific)
            history = save_json_history(json_file, json_data, last_wk[0], max_entries=weeks + 1)
            week_key = get_week_key(last_wk[0])
            print(f"  ✓ Saved week {week_key} ({len(history)} total entries)")
            
        except Exception as e:
            print(f"  ✗ Error processing week {week_num + 1}: {e}")
            continue
        
        # Move to previous week
        current_date = last_wk[0] - timedelta(days=7)
    
    print(f"\n✅ Backfill complete: {json_file}")
    print(f"   Total entries: {len(history)}")


def backfill_pipeline_velocity(start_date, weeks=12):
    """Backfill Pipeline Velocity JSON history from Salesforce"""
    print("\n" + "=" * 60)
    print("BACKFILLING PIPELINE VELOCITY JSON HISTORY")
    print("=" * 60)
    print(f"Start Date: {start_date}")
    print(f"Weeks: {weeks}")
    
    # Load template
    from openpyxl import load_workbook
    wb = load_workbook(pipe.TEMPLATE_FILE)
    ws = wb[pipe.SHEET_NAME]
    reps = pipe.load_template_and_reps()
    rep_names = [r[1] for r in reps]
    
    # Get report IDs
    rid_accts = pipe.report_id_from_note(ws, pipe.NOTE_ACCTS_OWNED)
    rid_open = pipe.report_id_from_note(ws, pipe.NOTE_OPEN_OPPS)
    rid_avg = pipe.report_id_from_note(ws, pipe.NOTE_AVG_DEAL)
    rid_age = pipe.report_id_from_note(ws, pipe.NOTE_AVG_OPP_AGE)
    
    print(f"\nReports: Accts={rid_accts}, Open={rid_open}, Avg={rid_avg}, Age={rid_age}")
    
    # Authenticate
    print("\nAuthenticating...")
    tok = pipe.get_jwt_token()
    print("✓ Connected")
    
    # Process each week
    script_dir = Path(__file__).parent
    json_file = script_dir / "pipeline_velocity_history.json"
    
    current_date = date.fromisoformat(start_date) if isinstance(start_date, str) else start_date
    
    print(f"\nProcessing {weeks} weeks starting from {current_date}...")
    
    for week_num in range(weeks):
        # Calculate week ranges for this iteration
        lw_s, lw_e = pipe.last_full_week(current_date)
        pw_s, pw_e = pipe.prior_full_week(current_date)
        
        print(f"\nWeek {week_num + 1}: {lw_s} to {lw_e}")
        
        try:
            
            # Fetch report data (these are current snapshots, not historical)
            # For historical accuracy, we'd need to use SOQL queries with date filters
            # For now, we'll use the current report data as a proxy
            rows_accts = pipe.analytics_report_data(tok, rid_accts)
            rows_open = pipe.analytics_report_data(tok, rid_open)
            rows_avg = pipe.analytics_report_data(tok, rid_avg)
            rows_age = pipe.analytics_report_data(tok, rid_age)
            
            accts_map = pipe.extract_accounts_or_open(rows_accts,
                ["# of Accounts Owned","Accounts Owned","Account Count","Accounts","Count","Record Count"])
            open_map = pipe.extract_accounts_or_open(rows_open,
                ["# of Open Opps","Open Opportunities","Open Opps","Open Opportunity Count","Count","Record Count"])
            avg_map = pipe.extract_avg_deal(rows_avg)
            age_map = pipe.extract_avg_opp_age(rows_age)
            
            # Win rate (YTD - doesn't change week to week)
            won_counts = pipe.won_counts_ytd_by_rep(tok)
            total_counts = pipe.total_counts_ytd_by_rep(tok)
            win_rate = {}
            for owner, total in total_counts.items():
                won = won_counts.get(owner, 0)
                win_rate[owner] = (won / total) if total else 0.0
            
            # Pipeline Created - use SOQL for this week
            lw_sum = pipe.pipeline_created_by_rep(tok, lw_s, lw_e)
            pw_sum = pipe.pipeline_created_by_rep(tok, pw_s, pw_e)
            
            # 6-week average
            trailing = pipe.six_trailing_weeks(current_date)
            six_sums = [pipe.pipeline_created_by_rep(tok, s, e) for (s, e) in trailing]
            owners = set().union(*[set(d.keys()) for d in six_sums]) if six_sums else set()
            six_avg = {o: (sum(d.get(o, 0.0) for d in six_sums) / len(six_sums)) if six_sums else 0.0
                       for o in owners}
            
            # Build JSON data
            json_data = {
                "period": {
                    "last_week": {"start": str(lw_s), "end": str(lw_e)},
                    "prior_week": {"start": str(pw_s), "end": str(pw_e)}
                },
                "data": {}
            }
            
            for rep in rep_names:
                lw = lw_sum.get(rep, 0.0)
                pw = pw_sum.get(rep, 0.0)
                avg = six_avg.get(rep, 0.0)
                wow = ((lw - pw) / pw * 100) if pw > 0 else 0
                vs_avg = ((lw - avg) / avg * 100) if avg > 0 else 0
                
                json_data["data"][rep] = {
                    "pipeline_created": {"last": lw, "prior": pw, "6w_avg": avg, 
                                         "wow_pct": wow, "vs_avg_pct": vs_avg},
                    "accounts_owned": accts_map.get(rep),
                    "open_opps": open_map.get(rep),
                    "win_rate": win_rate.get(rep, 0.0) * 100,
                    "avg_deal": avg_map.get(rep),
                    "avg_opp_age": age_map.get(rep)
                }
            
            # Save to history
            history = save_json_history(json_file, json_data, lw_s, max_entries=weeks + 1)
            week_key = get_week_key(lw_s)
            print(f"  ✓ Saved week {week_key} ({len(history)} total entries)")
            
        except Exception as e:
            print(f"  ✗ Error processing week {week_num + 1}: {e}")
            import traceback
            traceback.print_exc()
            continue
        
        # Move to previous week
        current_date = lw_s - timedelta(days=7)
    
    print(f"\n✅ Backfill complete: {json_file}")
    print(f"   Total entries: {len(history)}")


def main():
    parser = argparse.ArgumentParser(
        description="Backfill JSON history from Salesforce for Boca and Pipeline Velocity"
    )
    parser.add_argument(
        "--start-date",
        type=str,
        default="2025-11-03",
        help="Start date for backfill (YYYY-MM-DD). Default: 2025-11-03"
    )
    parser.add_argument(
        "--weeks",
        type=int,
        default=12,
        help="Number of weeks to backfill. Default: 12"
    )
    parser.add_argument(
        "--boca-only",
        action="store_true",
        help="Only backfill Boca Velocity"
    )
    parser.add_argument(
        "--pipeline-only",
        action="store_true",
        help="Only backfill Pipeline Velocity"
    )
    
    args = parser.parse_args()
    
    print("=" * 60)
    print("JSON HISTORY BACKFILL")
    print("=" * 60)
    
    if not args.pipeline_only:
        backfill_boca_velocity(args.start_date, args.weeks)
    
    if not args.boca_only:
        backfill_pipeline_velocity(args.start_date, args.weeks)
    
    print("\n" + "=" * 60)
    print("✅ ALL BACKFILLS COMPLETE")
    print("=" * 60)


if __name__ == "__main__":
    main()

