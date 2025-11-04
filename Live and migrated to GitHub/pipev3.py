import os
import re
import time
import jwt
import requests
from pathlib import Path
from datetime import datetime, timedelta, date
from collections import defaultdict
from openpyxl import load_workbook
from html_template_base import generate_html_document, generate_html_table, save_json_history, load_json_history

# =========================
# PATHS
# =========================
HOME = Path.home()
TEMPLATE_DIR  = HOME / "Desktop" / "My Weekly Reports - Python"
TEMPLATE_FILE = TEMPLATE_DIR / "PipelineTemplate.xlsx"
SHEET_NAME    = "Pipeline Velocity"

# Columns
COL_REP          = "A"
COL_LAST_WEEK    = "B"
COL_PRIOR_WEEK   = "C"
COL_WOW          = "D"
COL_SIX_WK_AVG   = "E"
COL_VS_AVG       = "F"
COL_ACCTS_OWNED  = "G"
COL_OPEN_OPPS    = "H"
COL_WIN_RATE     = "I"
COL_AVG_OPP_AGE  = "J"
COL_AVG_DEAL     = "K"

# Header cells with Notes (report links)
NOTE_ACCTS_OWNED = "G5"
NOTE_OPEN_OPPS   = "H5"
NOTE_WIN_RATE    = "I5"
NOTE_AVG_OPP_AGE = "J5"
NOTE_AVG_DEAL    = "K5"

# =========================
# SALESFORCE CONFIG - JWT Authentication
# =========================
# This script uses JWT (JSON Web Token) Bearer authentication to connect to Salesforce.
# 
# Required Configuration (in sf_config.py - NOT committed to GitHub):
# - SF_USERNAME: Your Salesforce username
# - SF_CONSUMER_KEY: Connected App Consumer Key
# - SF_DOMAIN: 'login' for production, 'test' for sandbox
# - PRIVATE_KEY_FILE: Path to your RSA private key file (.pem)
# 
# See sf_config.py.example for template and setup instructions.
# Security Note: sf_config.py is gitignored and will NOT be committed to GitHub
#
import sf_config

def get_jwt_token():
    """
    Generate JWT token for Salesforce authentication using RS256 algorithm.
    
    Returns:
        dict: Token response containing 'access_token' and 'instance_url'
    
    Raises:
        requests.HTTPError: If authentication fails
    """
    # Read the RSA private key file
    with open(sf_config.PRIVATE_KEY_FILE, "r") as f:
        private_key = f.read()
    
    # Build JWT claim (payload) with required Salesforce fields
    claim = {
        "iss": sf_config.SF_CONSUMER_KEY,  # Issuer: Your Connected App Consumer Key
        "sub": sf_config.SF_USERNAME,      # Subject: Your Salesforce username
        "aud": f"https://{sf_config.SF_DOMAIN}.salesforce.com",  # Audience: Salesforce login URL
        "exp": int(time.time()) + 300,      # Expiration: 5 minutes from now
    }
    
    # Encode the JWT token using RS256 algorithm (RSA signature)
    assertion = jwt.encode(claim, private_key, algorithm="RS256")
    
    # Exchange JWT assertion for access token
    token_url = f"https://{sf_config.SF_DOMAIN}.salesforce.com/services/oauth2/token"
    resp = requests.post(token_url, data={
        "grant_type": "urn:ietf:params:oauth:grant-type:jwt-bearer",  # JWT Bearer grant type
        "assertion": assertion,  # The JWT token we just created
    })
    resp.raise_for_status()
    return resp.json()  # Returns {access_token, instance_url, etc.}

def api_base(token, version="v61.0"):
    return f"{token['instance_url']}/services/data/{version}"

def sf_headers(token):
    return {"Authorization": f"Bearer {token['access_token']}", "Content-Type": "application/json"}

def soql_query(token, soql):
    url = f"{api_base(token)}/query"
    out = []
    r = requests.get(url, headers=sf_headers(token), params={"q": soql})
    r.raise_for_status()
    data = r.json()
    out.extend(data.get("records", []))
    while not data.get("done", True) and data.get("nextRecordsUrl"):
        r = requests.get(token["instance_url"] + data["nextRecordsUrl"], headers=sf_headers(token))
        r.raise_for_status()
        data = r.json()
        out.extend(data.get("records", []))
    return out

def analytics_report_data(token, report_id):
    """Fetch both detail and grouped rows into a list[dict]."""
    url = f"{api_base(token)}/analytics/reports/{report_id}?includeDetails=true"
    r = requests.get(url, headers=sf_headers(token))
    r.raise_for_status()
    j = r.json()

    # Detail rows first
    apis = j["reportMetadata"]["detailColumns"]
    info = j.get("reportExtendedMetadata", {}).get("detailColumnInfo", {})
    label_map = {a: info.get(a, {}).get("label", a) for a in apis}

    out = []
    for row in j.get("factMap", {}).get("T!T", {}).get("rows", []):
        rec = {}
        for api_name, cell in zip(apis, row["dataCells"]):
            rec[label_map[api_name]] = cell.get("value")
        out.append(rec)

    if out:
        return out

    # Grouped fallback
    agg_info = j.get("reportExtendedMetadata", {}).get("aggregateColumnInfo", {})
    down = j.get("groupingsDown", {})
    fact_map = j.get("factMap", {})

    for key, fact in fact_map.items():
        if key == "T!T":
            continue
        grouping_value = None
        if down and "groupings" in down:
            parts = key.split("!")
            if parts and parts[0].isdigit():
                idx = int(parts[0])
                if idx < len(down["groupings"]):
                    grouping_value = down["groupings"][idx].get("label")

        agg = fact.get("aggregates", [])
        if grouping_value and agg:
            rec = {"grouping": grouping_value}
            for i, a in enumerate(agg):
                label = agg_info.get(f"s{i}", {}).get("label") or f"Aggregate_{i}"
                rec[label] = a.get("value")
            out.append(rec)

    return out

# =========================
# NOTES / REPORT IDS
# =========================
def parse_report_id(text):
    m = re.search(r"/Report/(00O[a-zA-Z0-9]{12,17})/", text or "")
    return m.group(1) if m else None

def report_id_from_note(ws, cell_addr):
    c = ws[cell_addr]
    if not c.comment or not c.comment.text:
        raise RuntimeError(f"No Note found on {cell_addr}")
    rid = parse_report_id(c.comment.text)
    if not rid:
        raise RuntimeError(f"No report id found in Note at {cell_addr}")
    return rid

# =========================
# FIXED DATE LOGIC (Sunday→Saturday weeks)
# =========================
def last_full_week(ref_date=None):
    """
    Returns LAST WEEK per Salesforce definition:
    Sunday-Saturday of the previous calendar week.
    
    Example: If today is Monday Oct 28, 2025:
    - This week started Sunday Oct 26
    - LAST WEEK = Sunday Oct 19 to Saturday Oct 25
    """
    if ref_date is None:
        ref_date = date.today()
    
    # Python weekday: Monday=0, Tuesday=1, ..., Sunday=6
    dow = ref_date.weekday()
    
    # Find this week's Sunday
    if dow == 6:  # Today is Sunday
        this_weeks_sunday = ref_date
    else:
        # Days since last Sunday
        days_since_sunday = (dow + 1) % 7
        this_weeks_sunday = ref_date - timedelta(days=days_since_sunday)
    
    # LAST WEEK ends the day before this week's Sunday
    last_weeks_saturday = this_weeks_sunday - timedelta(days=1)
    last_weeks_sunday = last_weeks_saturday - timedelta(days=6)
    
    return last_weeks_sunday, last_weeks_saturday

def prior_full_week(ref_date=None):
    """Returns the week before last_full_week"""
    lw_s, lw_e = last_full_week(ref_date)
    pw_e = lw_s - timedelta(days=1)  # Day before last week's Sunday
    pw_s = pw_e - timedelta(days=6)
    return pw_s, pw_e

def six_trailing_weeks(ref_date=None):
    """Returns list of (start, end) tuples for the last 6 weeks including last_full_week"""
    out = []
    s, e = last_full_week(ref_date)
    out.append((s, e))
    for _ in range(5):
        e = s - timedelta(days=1)
        s = e - timedelta(days=6)
        out.append((s, e))
    return out

# =========================
# FIXED PIPELINE QUERY
# =========================
def pipeline_created_by_rep_manual_dates(token, start_d, end_d):
    """
    Query using explicit date range - for testing/debugging.
    """
    st = start_d.strftime("%Y-%m-%d")
    en = end_d.strftime("%Y-%m-%d")
    
    soql = f"""
      SELECT Professional_Services_Amount__c, Owner.Name, CreatedDate
      FROM Opportunity
      WHERE DAY_ONLY(CreatedDate) >= {st} 
            AND DAY_ONLY(CreatedDate) <= {en}
            AND Professional_Services_Amount__c != NULL
    """
    
    print(f"\n[DEBUG] Manual date query: {st} to {en}")
    recs = soql_query(token, soql)
    print(f"[DEBUG] Found {len(recs)} opportunities")
    
    sums = defaultdict(float)
    for r in recs:
        owner = (r.get("Owner") or {}).get("Name")
        amt = float(r.get("Professional_Services_Amount__c") or 0)
        if owner:
            sums[owner] += amt
    
    return sums

def pipeline_created_by_rep_salesforce_last_week(token):
    """
    Query using Salesforce's LAST_WEEK date literal - matches your report exactly.
    This is the SOURCE OF TRUTH for Column B.
    
    CRITICAL FILTERS (from your Salesforce report):
    - Excludes test accounts
    - Only includes opps CREATED BY specific users (not Owner!)
    - Groups/sums by Owner.Name for display
    """
    # These are the users from filter #8 - "Created By equals..."
    creator_names = [
        "AJ Rivera", "Chloe Lium", "Chris Fello", "Mike Meyer",
        "Michael Clinton", "Nick Errigo", "Taylor Copie"
    ]
    
    # Build the CreatedBy.Name filter
    creator_filter = ", ".join([f"'{name}'" for name in creator_names])
    
    soql = f"""
      SELECT Professional_Services_Amount__c, Owner.Name, CreatedBy.Name, CreatedDate, Account.Name
      FROM Opportunity
      WHERE CreatedDate = LAST_WEEK
            AND Professional_Services_Amount__c != NULL
            AND CreatedBy.Name IN ({creator_filter})
            AND (NOT Account.Name LIKE '%Test%')
            AND (NOT Account.Name LIKE '%test%')
            AND (NOT Account.Name LIKE '%ACME Corporation%')
    """
    
    print(f"\n[DEBUG] Salesforce LAST_WEEK query (filtered by Created By)")
    recs = soql_query(token, soql)
    print(f"[DEBUG] Found {len(recs)} opportunities")
    
    sums = defaultdict(float)
    for r in recs:
        # GROUP BY OWNER (not creator!) for the report display
        owner = (r.get("Owner") or {}).get("Name")
        creator = (r.get("CreatedBy") or {}).get("Name")
        amt = float(r.get("Professional_Services_Amount__c") or 0)
        created = r.get("CreatedDate", "")
        acct_name = (r.get("Account") or {}).get("Name", "")
        if owner:
            sums[owner] += amt
            # Show first 10 to verify
            if sum(1 for v in sums.values() if v != 0) <= 10:
                print(f"  - Owner: {owner} | Creator: {creator} | ${amt:,.0f} | Acct: {acct_name}")
    
    print(f"\n[DEBUG] Total pipeline: ${sum(sums.values()):,.0f}")
    return sums

def pipeline_created_by_rep(token, start_d, end_d):
    """
    Query opportunities created in date range [start_d, end_d] inclusive.
    Includes same filters as Salesforce report:
    - Excludes test accounts
    - Only includes opps CREATED BY specific users
    - Groups by Owner.Name
    """
    st = start_d.strftime("%Y-%m-%d")
    en = end_d.strftime("%Y-%m-%d")
    
    # These are the CREATORS from your Salesforce report filter #8
    creator_names = [
        "AJ Rivera", "Chloe Lium", "Chris Fello", "Mike Meyer",
        "Michael Clinton", "Nick Errigo", "Taylor Copie"
    ]
    creator_filter = ", ".join([f"'{name}'" for name in creator_names])
    
    soql = f"""
      SELECT Professional_Services_Amount__c, Owner.Name, CreatedBy.Name, CreatedDate
      FROM Opportunity
      WHERE DAY_ONLY(CreatedDate) >= {st} 
            AND DAY_ONLY(CreatedDate) <= {en}
            AND Professional_Services_Amount__c != NULL
            AND CreatedBy.Name IN ({creator_filter})
            AND (NOT Account.Name LIKE '%Test%')
            AND (NOT Account.Name LIKE '%test%')
            AND (NOT Account.Name LIKE '%ACME Corporation%')
    """
    
    print(f"\n[DEBUG] Querying pipeline from {st} to {en}")
    recs = soql_query(token, soql)
    print(f"[DEBUG] Found {len(recs)} opportunities")
    
    sums = defaultdict(float)
    for r in recs:
        # SUM BY OWNER (not creator)
        owner = (r.get("Owner") or {}).get("Name")
        amt = float(r.get("Professional_Services_Amount__c") or 0)
        if owner:
            sums[owner] += amt
            # Debug first few records
            if len([x for x in sums.values() if x > 0]) <= 3:
                creator = (r.get("CreatedBy") or {}).get("Name")
                print(f"  - Owner: {owner} | Creator: {creator} | ${amt:,.0f}")
    
    print(f"[DEBUG] Total: ${sum(sums.values()):,.0f}")
    return sums

def won_counts_ytd_by_rep(token):
    """Numerator for win rate."""
    yr = datetime.today().year
    st = f"{yr}-01-01"
    en = datetime.utcnow().strftime("%Y-%m-%d")
    soql = f"""
      SELECT Id, Owner.Name
      FROM Opportunity
      WHERE IsWon = true
        AND CloseDate >= {st} AND CloseDate <= {en}
    """
    recs = soql_query(token, soql)
    out = defaultdict(int)
    for r in recs:
        owner = (r.get("Owner") or {}).get("Name")
        if owner:
            out[owner] += 1
    return out

def total_counts_ytd_by_rep(token):
    """Denominator for win rate."""
    yr = datetime.today().year
    st = f"{yr}-01-01"
    en = datetime.utcnow().strftime("%Y-%m-%d")
    soql = f"""
      SELECT Id, Owner.Name
      FROM Opportunity
      WHERE CloseDate >= {st} AND CloseDate <= {en}
    """
    recs = soql_query(token, soql)
    out = defaultdict(int)
    for r in recs:
        owner = (r.get("Owner") or {}).get("Name")
        if owner:
            out[owner] += 1
    return out

# =========================
# METRIC EXTRACTORS (report rows)
# =========================
def extract_accounts_or_open(rows, metric_keys,
                             owner_keys=("Owner Name","Owner","Opportunity Owner","grouping")):
    """Generic extractor for accounts/open opps; picks the first matching label or numeric fallback."""
    def get_owner(r):
        for k in owner_keys:
            if k in r and r[k]:
                return str(r[k]).strip()
        return None

    def pick_value(r, keys):
        for k in keys:
            if k in r and r[k] is not None:
                try: return float(r[k])
                except Exception: pass
        # fallback: first numeric non-owner value
        for k, v in r.items():
            if k in owner_keys: 
                continue
            try:
                if v is not None and str(v).strip() != "":
                    return float(v)
            except Exception:
                continue
        return None

    out = {}
    for r in rows:
        owner = get_owner(r)
        if not owner:
            continue
        val = pick_value(r, metric_keys)
        if val is not None:
            out[owner] = val
    return out

def extract_avg_deal(rows, owner_keys=("Owner Name","Owner","Opportunity Owner","grouping")):
    """
    Prefer explicit average labels; else use Aggregate_1 from grouped results
    (matches your report samples).
    """
    avg_labels = ["Average Deal Size","Avg Deal Size","Average Amount","Avg Amount","Average"]
    def get_owner(r):
        for k in owner_keys:
            if k in r and r[k]:
                return str(r[k]).strip()
        return None

    out = {}
    for r in rows:
        owner = get_owner(r)
        if not owner:
            continue
        val = None
        for k in avg_labels:
            if k in r and r[k] is not None:
                try:
                    val = float(r[k]); break
                except Exception:
                    pass
        if val is None and "Aggregate_1" in r and r["Aggregate_1"] is not None:
            try:
                val = float(r["Aggregate_1"])
            except Exception:
                val = None
        if val is not None:
            out[owner] = val
    return out

def extract_avg_opp_age(rows, owner_keys=("Owner Name","Owner","Opportunity Owner","grouping")):
    """
    Extract Average Opportunity Age from Aggregate_2 column.
    """
    def get_owner(r):
        for k in owner_keys:
            if k in r and r[k]:
                return str(r[k]).strip()
        return None

    out = {}
    for r in rows:
        owner = get_owner(r)
        if not owner:
            continue
        
        if "Aggregate_2" in r and r["Aggregate_2"] is not None:
            try:
                out[owner] = float(r["Aggregate_2"])
            except Exception:
                pass
    
    return out

# =========================
# EXCEL IO
# =========================
def load_template_and_reps():
    if not TEMPLATE_FILE.exists():
        raise FileNotFoundError(f"Missing template: {TEMPLATE_FILE}")
    wb = load_workbook(TEMPLATE_FILE)
    ws = wb[SHEET_NAME]
    reps = []
    row = 6
    while True:
        v = ws[f"{COL_REP}{row}"].value
        if v is None:
            break
        if isinstance(v, str) and v.strip().lower().startswith("team average"):
            break
        reps.append((row, str(v).strip()))
        row += 1
    return wb, ws, reps

def write(ws, row, col, val):
    ws[f"{col}{row}"].value = val

def pct_change(new_val, old_val):
    if old_val in (None, 0) or new_val is None: 
        return None
    return (new_val - old_val) / old_val

# =========================
# MAIN
# =========================
def main():
    print("=" * 60)
    print("PIPELINE VELOCITY REPORT GENERATOR")
    print("=" * 60)
    
    # Show what weeks we're calculating
    today = date.today()
    lw_s, lw_e = last_full_week(today)
    pw_s, pw_e = prior_full_week(today)
    
    print(f"\nToday: {today.strftime('%A, %B %d, %Y')}")
    print(f"\nColumn B - Last Week: {lw_s} to {lw_e} ({(lw_e - lw_s).days + 1} days)")
    print(f"Column C - Prior Week: {pw_s} to {pw_e} ({(pw_e - pw_s).days + 1} days)")
    print("=" * 60)
    
    print("\nLoading template…")
    wb, ws, reps = load_template_and_reps()
    print(f"Found {len(reps)} reps: {[r[1] for r in reps]}")

    # Report IDs (Notes)
    rid_accts = report_id_from_note(ws, NOTE_ACCTS_OWNED)
    rid_open  = report_id_from_note(ws, NOTE_OPEN_OPPS)
    rid_avg   = report_id_from_note(ws, NOTE_AVG_DEAL)
    rid_age   = report_id_from_note(ws, NOTE_AVG_OPP_AGE)
    _rid_win  = report_id_from_note(ws, NOTE_WIN_RATE)

    # Auth
    print("\nAuthenticating with Salesforce...")
    token = get_jwt_token()
    print("✓ Connected")

    # Report data
    print("\nFetching report data...")
    rows_accts = analytics_report_data(token, rid_accts)
    rows_open  = analytics_report_data(token, rid_open)
    rows_avg   = analytics_report_data(token, rid_avg)
    rows_age   = analytics_report_data(token, rid_age)

    # Extract metrics
    accts_map = extract_accounts_or_open(rows_accts,
        ["# of Accounts Owned","Accounts Owned","Account Count","Accounts","Count","Record Count"])
    open_map  = extract_accounts_or_open(rows_open,
        ["# of Open Opps","Open Opportunities","Open Opps","Open Opportunity Count","Count","Record Count"])
    avg_map   = extract_avg_deal(rows_avg)
    age_map   = extract_avg_opp_age(rows_age)

    # Win rate from SOQL counts
    print("\nCalculating win rates...")
    won_counts   = won_counts_ytd_by_rep(token)
    total_counts = total_counts_ytd_by_rep(token)
    win_rate = {}
    for owner, total in total_counts.items():
        won = won_counts.get(owner, 0)
        win_rate[owner] = (won / total) if total else 0.0

    # Pipeline Created queries
    print("\n" + "=" * 60)
    print("QUERYING PIPELINE CREATED DATA")
    print("=" * 60)
    
    # Column B: Use Salesforce's LAST_WEEK (matches your report exactly)
    print("\n>>> COLUMN B: Using Salesforce LAST_WEEK")
    lw_sum = pipeline_created_by_rep_salesforce_last_week(token)
    
    # Also calculate what Python thinks last week is for comparison
    print(f"\n>>> For comparison, Python calculated last week as: {lw_s} to {lw_e}")
    lw_sum_python = pipeline_created_by_rep(token, lw_s, lw_e)
    
    print("\n>>> COMPARISON:")
    for rep_name in set(list(lw_sum.keys()) + list(lw_sum_python.keys())):
        sf_val = lw_sum.get(rep_name, 0)
        py_val = lw_sum_python.get(rep_name, 0)
        match = "✓" if sf_val == py_val else "✗ MISMATCH"
        print(f"  {rep_name}: SF=${sf_val:,.0f} | Python=${py_val:,.0f} {match}")
    
    # Column C: Prior week (week before LAST_WEEK)
    print(f"\n>>> COLUMN C: Prior week {pw_s} to {pw_e}")
    pw_sum = pipeline_created_by_rep(token, pw_s, pw_e)
    
    trailing = six_trailing_weeks(today)
    print(f"\nCalculating 6-week average from:")
    for i, (s, e) in enumerate(trailing, 1):
        print(f"  Week {i}: {s} to {e}")
    
    six_sums = [pipeline_created_by_rep(token, s, e) for (s, e) in trailing]
    owners = set().union(*[set(d.keys()) for d in six_sums]) if six_sums else set()
    six_avg = {o: (sum(d.get(o, 0.0) for d in six_sums) / len(six_sums)) if six_sums else 0.0
               for o in owners}

    # Prepare data for HTML and JSON
    dashboard_data = []
    json_data = {
        "period": {
            "last_week": {"start": str(lw_s), "end": str(lw_e)},
            "prior_week": {"start": str(pw_s), "end": str(pw_e)}
        },
        "data": {}
    }
    
    for row, rep in reps:
        # Gather all metrics
        lw = lw_sum.get(rep, 0.0)
        pw = pw_sum.get(rep, 0.0)
        avg = six_avg.get(rep, 0.0)
        wow = pct_change(lw, pw)
        vs_avg = pct_change(lw, avg)
        accts = accts_map.get(rep)
        open_opps = open_map.get(rep)
        wr = win_rate.get(rep, 0.0)
        avg_deal = avg_map.get(rep)
        avg_age = age_map.get(rep)
        
        # Store for JSON
        json_data["data"][rep] = {
            "pipeline_created": {"last": lw, "prior": pw, "6w_avg": avg, "wow_pct": wow * 100 if wow else 0, "vs_avg_pct": vs_avg * 100 if vs_avg else 0},
            "accounts_owned": accts,
            "open_opps": open_opps,
            "win_rate": wr * 100,
            "avg_deal": avg_deal,
            "avg_opp_age": avg_age
        }
        
        # Prepare HTML row
        dashboard_data.append({
            "rep": rep,
            "last_week": f"${lw:,.0f}",
            "prior_week": f"${pw:,.0f}",
            "6w_avg": f"${avg:,.0f}",
            "wow": f"{wow:+.1%}" if wow else "—",
            "vs_avg": f"{vs_avg:+.1%}" if vs_avg else "—",
            "accts": f"{accts:.0f}" if accts else "—",
            "open_opps": f"{open_opps:.0f}" if open_opps else "—",
            "win_rate": f"{wr:.1%}" if wr else "—",
            "avg_deal": f"${avg_deal:,.0f}" if avg_deal else "—",
            "avg_age": f"{avg_age:.0f}" if avg_age else "—"
        })
    
    # Write Excel (optional - keep for backward compatibility)
    print("\n" + "=" * 60)
    print("WRITING TO EXCEL")
    print("=" * 60)
    
    for row, rep in reps:
        lw  = lw_sum.get(rep, 0.0)
        pw  = pw_sum.get(rep, 0.0)
        avg = six_avg.get(rep, 0.0)

        write(ws, row, COL_LAST_WEEK,  lw)
        write(ws, row, COL_PRIOR_WEEK, pw)
        write(ws, row, COL_SIX_WK_AVG, avg)
        write(ws, row, COL_WOW,        pct_change(lw, pw))
        write(ws, row, COL_VS_AVG,     pct_change(lw, avg))

        write(ws, row, COL_ACCTS_OWNED, accts_map.get(rep))
        write(ws, row, COL_OPEN_OPPS,   open_map.get(rep))
        write(ws, row, COL_WIN_RATE,    win_rate.get(rep, 0.0))
        write(ws, row, COL_AVG_DEAL,    avg_map.get(rep))
        write(ws, row, COL_AVG_OPP_AGE, age_map.get(rep))

        # Console output
        print(f"\n{rep}:")
        print(f"  Col B (Last Week):  ${lw:>10,.0f}")
        print(f"  Col C (Prior Week): ${pw:>10,.0f}")
        print(f"  Col E (6-Wk Avg):   ${avg:>10,.0f}")
        print(f"  WoW: {pct_change(lw, pw):.1%}" if pct_change(lw, pw) else "  WoW: N/A")

    # Save Excel output
    excel_out = TEMPLATE_DIR / f"{datetime.today().strftime('%Y-%m-%d')}-Pipeline Template.xlsx"
    wb.save(excel_out)
    print(f"\n✅ Excel saved: {excel_out}")
    
    # Generate HTML Dashboard
    print("\n" + "=" * 60)
    print("GENERATING HTML DASHBOARD")
    print("=" * 60)
    
    # Define table columns
    columns = [
        {"key": "rep", "label": "Rep", "sortable": True},
        {"key": "last_week", "label": "Pipeline Created<br>Last Week", "sortable": True, "align": "right"},
        {"key": "prior_week", "label": "Pipeline Created<br>Prior Week", "sortable": True, "align": "right"},
        {"key": "6w_avg", "label": "Pipeline Created<br>6W Avg", "sortable": True, "align": "right"},
        {"key": "wow", "label": "WoW", "sortable": True, "align": "right"},
        {"key": "vs_avg", "label": "Vs Avg", "sortable": True, "align": "right"},
        {"key": "accts", "label": "Accts<br>Owned", "sortable": True, "align": "right"},
        {"key": "open_opps", "label": "Open<br>Opps", "sortable": True, "align": "right"},
        {"key": "win_rate", "label": "Win<br>Rate", "sortable": True, "align": "right"},
        {"key": "avg_deal", "label": "Avg<br>Deal", "sortable": True, "align": "right"},
        {"key": "avg_age", "label": "Avg Opp<br>Age", "sortable": True, "align": "right"}
    ]
    
    # Generate HTML table
    table_html = generate_html_table(columns, dashboard_data)
    
    # Generate complete HTML document
    title = "Pipeline Velocity Report"
    subtitle = f"Pipeline Velocity - Last Week: {lw_s} to {lw_e} | Prior Week: {pw_s} to {pw_e}"
    
    # Add JavaScript for sorting
    js_script = """
<script>
(function() {
  const tbody = document.querySelector('tbody');
  const headers = document.querySelectorAll('th.sort');
  let sortKey = 'rep';
  let sortDir = 'asc';
  
  function render() {
    const rows = Array.from(tbody.querySelectorAll('tr'));
    rows.sort((a, b) => {
      const aVal = a.querySelector(`td:nth-child(${getColumnIndex(sortKey)})`).textContent.trim();
      const bVal = b.querySelector(`td:nth-child(${getColumnIndex(sortKey)})`).textContent.trim();
      const aNum = parseFloat(aVal.replace(/[^0-9.-]/g, ''));
      const bNum = parseFloat(bVal.replace(/[^0-9.-]/g, ''));
      if (!isNaN(aNum) && !isNaN(bNum)) {
        return (aNum - bNum) * (sortDir === 'asc' ? 1 : -1);
      }
      return aVal.localeCompare(bVal) * (sortDir === 'asc' ? 1 : -1);
    });
    rows.forEach(r => tbody.appendChild(r));
  }
  
  function getColumnIndex(key) {
    const cols = ['rep', 'last_week', 'prior_week', '6w_avg', 'wow', 'vs_avg', 
                  'accts', 'open_opps', 'win_rate', 'avg_deal', 'avg_age'];
    return cols.indexOf(key) + 1;
  }
  
  headers.forEach(th => {
    th.addEventListener('click', () => {
      sortKey = th.dataset.key;
      sortDir = sortDir === 'asc' ? 'desc' : 'asc';
      render();
    });
  });
})();
</script>
"""
    
    html_content = generate_html_document(title, subtitle, table_html, js_script)
    
    # Save HTML (in script directory)
    script_dir = Path(__file__).parent
    html_file = script_dir / "pipeline_velocity_latest.html"
    with open(html_file, 'w', encoding='utf-8') as f:
        f.write(html_content)
    print(f"✅ HTML Dashboard saved: {html_file}")
    
    # Save JSON History (week-specific: uses lw_s as period start date)
    json_file = script_dir / "pipeline_velocity_history.json"
    history = save_json_history(json_file, json_data, lw_s, max_entries=12)
    week_key = lw_s.strftime("%Y-W%V")
    print(f"✅ JSON History saved: {json_file} ({len(history)} entries) - Week: {week_key}")
    
    print("\n" + "=" * 60)
    print("✅ COMPLETE")
    print("=" * 60)
    print(f"\n📊 Outputs:")
    print(f"   Excel: {excel_out}")
    print(f"   HTML:  {html_file}")
    print(f"   JSON:  {json_file}")
    print(f"\n🌐 GitHub Pages URL:")
    print(f"   https://AstonAtInnovativeSol.github.io/salesforce-flow-metrics/pipeline_velocity_latest.html")

if __name__ == "__main__":
    main()
